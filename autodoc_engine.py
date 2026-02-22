"""
AutoDoc - Engineering Report Automation Pipeline
Core automation engine: reads Excel component data, generates PDF reports
"""

import os
import json
import csv
import datetime
import openpyxl
from jinja2 import Environment, FileSystemLoader
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


# ── Config ────────────────────────────────────────────────────────────────────
LOG_FILE = "report_log.csv"
OUTPUT_DIR = "generated_reports"
TEMPLATE_DIR = "templates"

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(TEMPLATE_DIR, exist_ok=True)


# ── Excel Parser ──────────────────────────────────────────────────────────────
def parse_excel(filepath: str) -> list[dict]:
    """
    Reads a component spec Excel file.
    Expected columns: Component ID, Name, Type, Voltage Rating (V),
                      Current Rating (A), Material, Status, Engineer, Notes
    Returns a list of component dicts.
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [str(cell.value).strip() for cell in ws[1]]
    components = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            comp = dict(zip(headers, row))
            # Clean up None values
            comp = {k: (v if v is not None else "N/A") for k, v in comp.items()}
            components.append(comp)

    wb.close()
    return components


# ── PDF Generator ─────────────────────────────────────────────────────────────
BRAND_BLUE   = colors.HexColor("#1A3A5C")
BRAND_ORANGE = colors.HexColor("#F57C00")
LIGHT_GRAY   = colors.HexColor("#F5F5F5")
MID_GRAY     = colors.HexColor("#9E9E9E")
DARK_TEXT    = colors.HexColor("#212121")


def build_styles():
    styles = getSampleStyleSheet()
    custom = {
        "Title": ParagraphStyle("Title", fontName="Helvetica-Bold",
                                fontSize=22, textColor=BRAND_BLUE,
                                spaceAfter=4, alignment=TA_LEFT),
        "Subtitle": ParagraphStyle("Subtitle", fontName="Helvetica",
                                   fontSize=11, textColor=MID_GRAY,
                                   spaceAfter=16),
        "SectionHeader": ParagraphStyle("SectionHeader", fontName="Helvetica-Bold",
                                        fontSize=12, textColor=BRAND_BLUE,
                                        spaceBefore=14, spaceAfter=6),
        "Body": ParagraphStyle("Body", fontName="Helvetica",
                               fontSize=9, textColor=DARK_TEXT,
                               spaceAfter=4, leading=14),
        "Footer": ParagraphStyle("Footer", fontName="Helvetica",
                                 fontSize=8, textColor=MID_GRAY,
                                 alignment=TA_CENTER),
    }
    return custom


def generate_pdf_report(components: list[dict], project_name: str,
                         engineer: str = "AutoDoc System") -> str:
    """
    Generates a formatted PDF engineering report from component data.
    Returns the output filepath.
    """
    timestamp   = datetime.datetime.now()
    report_id   = f"RPT-{timestamp.strftime('%Y%m%d-%H%M%S')}"
    filename    = f"{OUTPUT_DIR}/{report_id}_{project_name.replace(' ', '_')}.pdf"
    generated_at = timestamp.strftime("%B %d, %Y  %H:%M:%S")

    doc = SimpleDocTemplate(filename, pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)

    styles = build_styles()
    story  = []

    # ── Header ────────────────────────────────────────────────────────────────
    story.append(Paragraph("AutoDoc", styles["Title"]))
    story.append(Paragraph(f"Engineering Component Specification Report", styles["Subtitle"]))
    story.append(HRFlowable(width="100%", thickness=2, color=BRAND_ORANGE, spaceAfter=12))

    # Meta table
    meta_data = [
        ["Report ID",    report_id,         "Project",   project_name],
        ["Generated",    generated_at,      "Engineer",  engineer],
        ["Components",   str(len(components)), "Status", "DRAFT"],
    ]
    meta_table = Table(meta_data, colWidths=[1.1*inch, 2.4*inch, 1*inch, 2.4*inch])
    meta_table.setStyle(TableStyle([
        ("FONTNAME",    (0,0), (-1,-1), "Helvetica"),
        ("FONTNAME",    (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME",    (2,0), (2,-1), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("TEXTCOLOR",   (0,0), (0,-1), BRAND_BLUE),
        ("TEXTCOLOR",   (2,0), (2,-1), BRAND_BLUE),
        ("BACKGROUND",  (0,0), (-1,-1), LIGHT_GRAY),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.white),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [LIGHT_GRAY, colors.white]),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 16))

    # ── Component Table ────────────────────────────────────────────────────────
    story.append(Paragraph("Component Specifications", styles["SectionHeader"]))

    # Determine columns from data
    if components:
        col_keys = list(components[0].keys())
        header_row = col_keys
        data_rows  = [[str(c.get(k, "")) for k in col_keys] for c in components]
        table_data = [header_row] + data_rows

        col_count  = len(col_keys)
        col_width  = (7 * inch) / col_count
        col_widths = [col_width] * col_count

        comp_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        comp_table.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), BRAND_BLUE),
            ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 8),
            ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, LIGHT_GRAY]),
            ("GRID",         (0,0), (-1,-1), 0.3, MID_GRAY),
            ("TOPPADDING",   (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ("LEFTPADDING",  (0,0), (-1,-1), 6),
            ("ALIGN",        (0,0), (-1,-1), "LEFT"),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(comp_table)

    story.append(Spacer(1, 16))

    # ── Summary ────────────────────────────────────────────────────────────────
    story.append(Paragraph("Summary", styles["SectionHeader"]))

    statuses = {}
    types    = {}
    for c in components:
        s = str(c.get("Status", "Unknown"))
        t = str(c.get("Type",   "Unknown"))
        statuses[s] = statuses.get(s, 0) + 1
        types[t]    = types.get(t, 0) + 1

    summary_items = [
        ["Total Components", str(len(components))],
        ["Approved",         str(statuses.get("Approved", 0))],
        ["Under Review",     str(statuses.get("Under Review", 0))],
        ["Pending",          str(statuses.get("Pending", 0))],
    ]
    for t, cnt in types.items():
        summary_items.append([f"Type: {t}", str(cnt)])

    sum_table = Table(summary_items, colWidths=[2*inch, 1.5*inch])
    sum_table.setStyle(TableStyle([
        ("FONTNAME",    (0,0), (-1,-1), "Helvetica"),
        ("FONTNAME",    (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("TEXTCOLOR",   (0,0), (0,-1), BRAND_BLUE),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[LIGHT_GRAY, colors.white]),
        ("GRID",        (0,0), (-1,-1), 0.3, colors.white),
        ("TOPPADDING",  (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
    ]))
    story.append(sum_table)
    story.append(Spacer(1, 20))

    # ── Footer ─────────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=1, color=MID_GRAY, spaceAfter=6))
    story.append(Paragraph(
        f"Generated by AutoDoc Automation Pipeline · {generated_at} · {report_id}  |  CONFIDENTIAL — INTERNAL USE ONLY",
        styles["Footer"]
    ))

    doc.build(story)
    return filename


# ── Logger ────────────────────────────────────────────────────────────────────
def log_report(report_id: str, project: str, component_count: int,
               filepath: str, status: str = "Success"):
    """Appends a record to the CSV report log (feeds Power BI)."""
    file_exists = os.path.isfile(LOG_FILE)
    with open(LOG_FILE, "a", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["report_id", "project", "component_count",
                             "filepath", "status", "generated_at"])
        writer.writerow([
            report_id, project, component_count,
            filepath, status,
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])


# ── Main Pipeline ─────────────────────────────────────────────────────────────
def run_pipeline(excel_path: str, project_name: str, engineer: str = "AutoDoc") -> dict:
    """
    Full pipeline: parse Excel → generate PDF → log result.
    Returns a result dict.
    """
    result = {"status": "error", "message": "", "filepath": "", "report_id": ""}
    try:
        print(f"[AutoDoc] Parsing: {excel_path}")
        components = parse_excel(excel_path)
        print(f"[AutoDoc] Found {len(components)} components")

        print("[AutoDoc] Generating PDF report...")
        pdf_path = generate_pdf_report(components, project_name, engineer)
        report_id = os.path.basename(pdf_path).split("_")[0]

        print("[AutoDoc] Logging result...")
        log_report(report_id, project_name, len(components), pdf_path)

        result = {
            "status":          "success",
            "message":         f"Report generated: {pdf_path}",
            "filepath":        pdf_path,
            "report_id":       report_id,
            "component_count": len(components),
        }
        print(f"[AutoDoc] Done → {pdf_path}")

    except Exception as e:
        result["message"] = str(e)
        print(f"[AutoDoc] ERROR: {e}")

    return result


if __name__ == "__main__":
    # Quick smoke-test with sample data
    import sys
    if len(sys.argv) >= 3:
        r = run_pipeline(sys.argv[1], sys.argv[2],
                         sys.argv[3] if len(sys.argv) > 3 else "AutoDoc")
        print(json.dumps(r, indent=2))
    else:
        print("Usage: python autodoc_engine.py <excel_file> <project_name> [engineer]")
